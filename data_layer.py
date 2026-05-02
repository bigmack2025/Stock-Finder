"""
Data layer: reads the master xlsx, builds a clean DataFrame with engineered features,
and saves to data/current_universe.parquet (date-stamped).

Run once after dropping a new xlsx in this folder:
    python data_layer.py

Council reco fixes applied:
  - International sheet folded in with `region` column (P1)
  - Date-stamped parquet versioning, last 3 kept (P2)
  - Source provenance and load timestamp persisted alongside data (P2)
"""

from __future__ import annotations

import json
import math
import re
import shutil
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).parent
DATA_DIR = ROOT / "data"
DATA_DIR.mkdir(exist_ok=True)
ARCHIVE_DIR = DATA_DIR / "archive"
ARCHIVE_DIR.mkdir(exist_ok=True)

XLSX_DEFAULT = ROOT / "Public Biotech Stocks Master List.xlsx"

MODALITIES = [
    "Oncology",
    "Gene/Cell Therapy",
    "RNA/Antisense",
    "Neurology",
    "Immunology",
    "Rare Disease",
    "Metabolic",
    "Cardiovascular",
    "Ophthalmology",
    "Vaccines",
    "Dermatology",
    "AI/Platform",
    "Psychedelics/CNS",
    "Biosimilars",
    "General Biotech",
]


# ---------------------------------------------------------------------------
# Modality parsing
# ---------------------------------------------------------------------------

# For the International sheet's "Notes / Focus" free-text column, we keyword-match
# against a curated lexicon. This is intentionally conservative — better to land
# on "General Biotech" than mislabel.
INTL_KEYWORDS = {
    "Oncology": ["oncology", "tumor", "cancer", "PD-1", "pd-l1", "ADC", "her2", "egfr"],
    "Gene/Cell Therapy": ["gene therapy", "cell therapy", "CAR-T", "car-t", "AAV", "iPSC"],
    "RNA/Antisense": ["mRNA", "RNAi", "siRNA", "antisense"],
    "Neurology": ["neuro", "alzheimer", "parkinson", "ALS", "epilepsy", "CNS"],
    "Immunology": ["immunology", "autoimmune", "lupus", "psoriasis", "ulcerative"],
    "Rare Disease": ["rare disease", "orphan"],
    "Metabolic": ["metabolic", "obesity", "diabet", "GLP-1", "NASH"],
    "Cardiovascular": ["cardio", "heart"],
    "Ophthalmology": ["ophthalm", "retina", "eye"],
    "Vaccines": ["vaccine"],
    "Dermatology": ["derm", "skin"],
    "Biosimilars": ["biosimilar"],
    "AI/Platform": ["AI", "platform", "discovery engine"],
    "Psychedelics/CNS": ["psychedelic"],
}


def parse_subsector(subsector: str | None) -> list[str]:
    """Split a US-sheet sub-sector cell into the canonical modality list."""
    if not subsector:
        return ["General Biotech"]
    parts = [p.strip() for p in subsector.split(",")]
    out: list[str] = []
    for p in parts:
        if "Metabolic" in p:
            out.append("Metabolic")
        elif "Rare" in p:
            out.append("Rare Disease")
        else:
            out.append(p)
    seen = set()
    deduped = []
    for m in out:
        if m not in seen:
            deduped.append(m)
            seen.add(m)
    return deduped


def parse_intl_notes(notes: str | None) -> list[str]:
    """Keyword-match the International 'Notes / Focus' free-text column."""
    if not notes:
        return ["General Biotech"]
    matches: list[str] = []
    text_lc = notes.lower()
    for mod, keywords in INTL_KEYWORDS.items():
        for kw in keywords:
            if kw.lower() in text_lc:
                matches.append(mod)
                break
    if not matches:
        return ["General Biotech"]
    seen = set()
    return [m for m in matches if not (m in seen or seen.add(m))]


def size_band(mkt_cap_m: float | None) -> str:
    if mkt_cap_m is None or pd.isna(mkt_cap_m) or mkt_cap_m <= 0:
        return "unknown"
    if mkt_cap_m < 300:
        return "micro"
    if mkt_cap_m < 2_000:
        return "small"
    if mkt_cap_m < 10_000:
        return "mid"
    if mkt_cap_m < 50_000:
        return "large"
    return "mega"


SIZE_BAND_ORDINAL = {"unknown": 0, "micro": 1, "small": 2, "mid": 3, "large": 4, "mega": 5}


# ---------------------------------------------------------------------------
# Load
# ---------------------------------------------------------------------------

def _load_us_biotech(wb) -> list[dict]:
    rows: list[dict] = []
    ws = wb["US Biotech"]
    headers = [c.value for c in ws[1]]
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not raw[0]:
            continue
        rec = dict(zip(headers, raw))
        rows.append(
            {
                "ticker": str(rec["Ticker"]).strip(),
                "name": rec["Company Name"],
                "mkt_cap_m": rec["Mkt Cap ($M)"],
                "revenue_m": rec["Revenue ($M)"],
                "stage_raw": rec["Stage"],
                "subsector_raw": rec["Sub-sector (heuristic)"],
                "industry": rec["Industry"],
                "region": "US",
                "exchange_hint": None,
                "notes": None,
                "source_sheet": "US Biotech",
            }
        )
    return rows


def _load_international(wb) -> list[dict]:
    rows: list[dict] = []
    if "International" not in wb.sheetnames:
        return rows
    ws = wb["International"]
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not raw[0]:
            continue
        ticker = str(raw[0]).strip()
        rows.append(
            {
                "ticker": ticker,
                "name": raw[1],
                "mkt_cap_m": None,           # filled by yfinance later
                "revenue_m": None,
                "stage_raw": None,
                "subsector_raw": None,
                "industry": "Biotechnology (Intl)",
                "region": str(raw[2]) if raw[2] else "Intl",
                "exchange_hint": str(raw[2]) if raw[2] else None,
                "notes": raw[3],
                "source_sheet": "International",
            }
        )
    return rows


def load_universe(xlsx_path: Path = XLSX_DEFAULT) -> pd.DataFrame:
    """Read US Biotech + International sheets, normalize, return DataFrame."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    rows = _load_us_biotech(wb) + _load_international(wb)
    df = pd.DataFrame(rows)

    # Numerics
    df["mkt_cap_m"] = pd.to_numeric(df["mkt_cap_m"], errors="coerce")
    df["revenue_m"] = pd.to_numeric(df["revenue_m"], errors="coerce").fillna(0)

    # Logged features
    df["log_mkt_cap"] = df["mkt_cap_m"].apply(
        lambda x: math.log10(x) if pd.notna(x) and x > 0 else None
    )
    df["log_revenue"] = df["revenue_m"].apply(lambda x: math.log10(x + 1) if x >= 0 else 0)
    df["has_revenue"] = (df["revenue_m"] > 0).astype(int)
    df["size_band"] = df["mkt_cap_m"].apply(size_band)
    df["size_band_ord"] = df["size_band"].map(SIZE_BAND_ORDINAL)

    # Modalities — branch by source sheet
    def _mods(row):
        if row["source_sheet"] == "US Biotech":
            return parse_subsector(row["subsector_raw"])
        return parse_intl_notes(row["notes"])

    df["modalities"] = df.apply(_mods, axis=1)

    # ---- Pipeline-extractor enrichment (M5 shipped, partial coverage) ----
    # If we have a cached rich-modality extraction for this ticker (from
    # `pipeline_extractor.py` reading the company's actual 10-K Item 1), we
    # surface the richer tags + therapeutic areas in *parallel* columns:
    #     `rich_modalities`        — list of drug-class tags
    #     `rich_therapeutic_areas` — list of TA tags
    #     `pipeline_filed`         — date of the source 10-K
    #     `modality_source`        — "lexicon" / "llm" / "xlsx" (fallback)
    #
    # IMPORTANT: We do NOT mutate `modalities` itself (which drives the
    # similarity engine's Jaccard calc). Until we have rich extractions for the
    # ENTIRE universe, mixing rich tags with xlsx tags creates an asymmetry —
    # the rich-tagged names overlap on more dimensions and get unfair similarity
    # advantage. The richer info is *displayed* in the UI but doesn't drive
    # ranking until M6 universe-wide extraction.
    pipeline_dir = ROOT / "data" / "pipeline"
    rich_mods_col: list[list[str]] = []
    rich_tas_col: list[list[str]] = []
    pipeline_filed_col: list[str | None] = []
    modality_source_col: list[str] = []
    rich_used = 0
    for _, row in df.iterrows():
        tk = row["ticker"]
        cache_path = pipeline_dir / f"{tk}.json" if pipeline_dir.exists() else None
        if cache_path and cache_path.exists():
            try:
                payload = json.loads(cache_path.read_text())
                mods = payload.get("modalities") or []
                tas = payload.get("therapeutic_areas") or []
                if mods or tas:
                    rich_mods_col.append(mods)
                    rich_tas_col.append(tas)
                    pipeline_filed_col.append(payload.get("filing_filed"))
                    modality_source_col.append(payload.get("method", "lexicon"))
                    rich_used += 1
                    continue
            except Exception:
                pass
        rich_mods_col.append([])
        rich_tas_col.append([])
        pipeline_filed_col.append(None)
        modality_source_col.append("xlsx")
    df["rich_modalities"] = rich_mods_col
    df["rich_therapeutic_areas"] = rich_tas_col
    df["pipeline_filed"] = pipeline_filed_col
    df["modality_source"] = modality_source_col

    # ---- combined_modalities — what the similarity Jaccard actually uses ----
    # Strategy: when rich tags are available (M6 universe-wide extraction
    # complete), use them (modality + therapeutic area combined as Jaccard
    # input). When not available, fall back to xlsx tags. This keeps the
    # Jaccard input non-empty for every ticker even if rich coverage is partial.
    combined = []
    for i, row in df.iterrows():
        rich_mods = list(row["rich_modalities"] or [])
        rich_tas = list(row["rich_therapeutic_areas"] or [])
        xlsx_mods = list(row["modalities"] or [])
        if rich_mods or rich_tas:
            # Rich path — use the union of modality + TA + xlsx (xlsx as coarse fallback dim)
            combo = rich_mods + rich_tas + xlsx_mods
        else:
            combo = xlsx_mods
        # De-dup, preserve order
        seen = set()
        deduped = [m for m in combo if not (m in seen or seen.add(m))]
        combined.append(deduped)
    df["combined_modalities"] = combined

    df["primary_modality"] = df["modalities"].apply(lambda lst: lst[0] if lst else "General Biotech")
    df["is_compound_modality"] = df["modalities"].apply(lambda lst: int(len(lst) > 1))

    # Note: the one-hot mod_* columns continue to use the coarse MODALITIES
    # constant for stable column shape. Rich modalities live in the `modalities`
    # list column and drive the Jaccard similarity calc directly.
    for m in MODALITIES:
        df[f"mod_{m}"] = df["modalities"].apply(lambda lst, m=m: int(m in lst))

    if rich_used:
        print(f"  pipeline override: {rich_used} tickers using rich modalities from 10-K extraction")

    return df


# ---------------------------------------------------------------------------
# Save (date-stamped)
# ---------------------------------------------------------------------------

def save(df: pd.DataFrame, keep_last: int = 3) -> Path:
    """Save current_universe.parquet plus a date-stamped archive copy."""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    # Authoritative current copy
    current = DATA_DIR / "current_universe.parquet"
    df.to_parquet(current, index=False)
    df.drop(columns=["modalities"]).to_csv(DATA_DIR / "current_universe.csv", index=False)

    # Archive copy
    archive = ARCHIVE_DIR / f"universe_{today}.parquet"
    shutil.copy(current, archive)

    # Trim archive to last N
    archives = sorted(ARCHIVE_DIR.glob("universe_*.parquet"))
    for old in archives[:-keep_last]:
        old.unlink()

    meta = {
        "loaded_at_utc": datetime.now(timezone.utc).isoformat(),
        "n_companies": len(df),
        "n_us": int((df["region"] == "US").sum()),
        "n_intl": int((df["region"] != "US").sum()),
        "n_with_mkt_cap": int(df["mkt_cap_m"].notna().sum()),
        "n_with_revenue": int(df["has_revenue"].sum()),
        "modalities": MODALITIES,
        "size_band_ordinal": SIZE_BAND_ORDINAL,
        "source_xlsx": str(XLSX_DEFAULT.name),
    }
    (DATA_DIR / "meta.json").write_text(json.dumps(meta, indent=2))
    return current


if __name__ == "__main__":
    df = load_universe()
    out = save(df)
    n_us = (df["region"] == "US").sum()
    n_intl = (df["region"] != "US").sum()
    print(f"Saved {len(df)} companies ({n_us} US, {n_intl} international) → {out}")
    print(df[["ticker", "name", "region", "mkt_cap_m", "primary_modality"]].head(5).to_string(index=False))
    print("...")
    print(df[df["region"] != "US"][["ticker", "name", "region", "primary_modality"]].head(5).to_string(index=False))
